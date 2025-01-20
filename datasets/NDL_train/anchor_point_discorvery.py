import xlrd
import xlwt
import openpyxl
from openpyxl import load_workbook
import torch
from scipy import fft
import numpy as np
from KDEpy import FFTKDE
import matplotlib.pyplot as plt
import matplotlib
from math import pow
import random
import tsai
from tsai.all import *
from tsai.inference import load_learner
from sklearn.metrics import roc_curve
from sklearn.metrics import auc
from sklearn.metrics import recall_score, accuracy_score
from sklearn.metrics import precision_score, f1_score
import time
import copy
import torch.optim as optim
from torch.autograd import Variable
import math
from sklearn.cluster import KMeans
import pandas as pd
from sklearn.metrics import mean_squared_error
import sklearn.linear_model
import platform
import pathlib

TRUE_NUM = 12
TIME_STEP = 10
# number_id = [4, 6, 8, 10, 12, 14]
number_id = [4]


def data_normal(orign_data):
    ff = orign_data.numpy()
    ff = list(map(abs, ff))
    ss = abs(np.argmax(np.bincount(ff[-50:])) - np.argmax(np.bincount(ff[0:50])))
    result = []
    if ss > 5:
        for j in range(0, len(ff), 128):
            if j >= 256:
                ff1 = ff[j - 256:j]
                Fs = 2  # 采样频率 按香农采样定理，采样频率为最高频率的2倍即可
                T = 1 / Fs  # 采样周期
                L = len(ff1)  # 信号长度
                t = [i * T for i in range(L)]
                t = np.array(t)
                complex_array = fft.fft(ff1)  # 快速傅里叶变换，返回结果为1000个复数，复数的模为幅值；复数的角度为相频
                # 得到分解波的频率序列
                freqs = fft.fftfreq(t.size, t[1] - t[0])  # 从0开始,既有正的频率，又有负的，只取一半即可
                # 复数的模为信号的振幅（能量大小）
                pows = np.abs(complex_array) / (Fs * 2)
                pows[freqs == 0] = pows[freqs == 0] / 2  # 频率为0的重复累积了两次，需要减半处理
                pows = pows[freqs >= 0]  # 取出对应的正频率的幅值
                # freqs = freqs[freqs >= 0]
                # plt.title('FFT')
                # plt.xlabel('Frequency')
                # plt.ylabel('Power')
                # plt.tick_params(labelsize=10)
                # plt.grid(linestyle=':')
                # plt.plot(freqs[freqs >= 0], pows[freqs >= 0], c='orangered', label='Frequency')
                # plt.legend()
                # plt.tight_layout()
                # plt.show()
                result.append(pows[0])

        xx = np.zeros(len(result))
        Y = np.array(list(zip(xx, result)))
        kmeans = KMeans(n_clusters=2).fit(Y)
        pred = kmeans.predict(Y)

        # print(result)
        # print(pred)
        local = []
        for key, iter in itertools.groupby(pred):
            local.append((key, len(list(iter))))
        # print(local)
        sum = len(local) - 1
        j = 0
        while (j <= sum):
            if local[j][1] <= 2:
                if j == 0:
                    yuanzu = (local[j + 1][0], local[j][1] + local[j + 1][1])
                    local[j + 1] = yuanzu
                    del (local[j])
                if j == len(local) - 1:
                    yuanzu = (local[j - 1][0], local[j - 1][1] + local[j][1])
                    local[j - 1] = yuanzu
                    del (local[j])
                if j > 0 and j < len(local) - 1:
                    yuanzu = (local[j - 1][0], local[j - 1][1] + local[j][1] + local[j + 1][1])
                    local[j - 1] = yuanzu
                    del (local[j + 1])
                    del (local[j])
                    j -= 1
            sum = len(local) - 1
            j += 1
        # print(local)
        if len(local) >= 2:
            loc = 256 + 128 * (local[0][1])
            f_min = ff[0]
            counts = np.bincount(ff[loc:])
            s_min = np.argmax(counts)
            norm = ff
            for i in range(loc):
                norm[i] = ff[i] - f_min
            for i in range(loc, len(ff)):
                if ff[i] == 82:
                    norm[i] = ff[i] - f_min
                else:
                    norm[i] = ff[i] - s_min
            norm_data = torch.tensor(norm, dtype=torch.float)
        else:
            for i in range(len(orign_data)):
                orign_data[i] = torch.abs(orign_data[i])
            d_min = orign_data[0]
            norm_data = orign_data - d_min
    else:
        for i in range(len(orign_data)):
            orign_data[i] = torch.abs(orign_data[i])
        d_min = orign_data[0]
        norm_data = orign_data - d_min
    return norm_data


def load_data(data_file):
    wb_node = load_workbook(data_file)
    sh_node = wb_node['Sheet']
    data, label, postion = [], [], []
    rssi, etx, succ, pos, ff, tr = [], [], [], [], [], []
    rssi1, etx1, ff1, ff2, ff3, succ1, label1, tr1 = [], [], [], [], [], [], [], []
    data_label = []
    data_all = []
    lens = sh_node.max_row

    for i in range(1, lens + 1):
        if int(sh_node.cell(i, 3).value) == -100:
            rssi.append(-82)
        else:
            rssi.append(int(sh_node.cell(i, 3).value))
        etx.append(sh_node.cell(i, 4).value)
        postion.append((int(sh_node.cell(i, 5).value), int(sh_node.cell(i, 6).value)))
        succ.append(int(sh_node.cell(i, 7).value))
        data_label.append([int(sh_node.cell(i, 8).value)])
        if i == 1:
            tr.append(0)
        else:
            if rssi[i - 1] > rssi[i - 2]:
                tr.append(1)
            else:
                if rssi[i - 1] < rssi[i - 2]:
                    tr.append(1)
                else:
                    tr.append(0)

    # ff = fft.irfft(ff)
    rssi = torch.tensor(rssi, dtype=torch.float)
    etx = torch.tensor(etx, dtype=torch.float)
    succ = torch.tensor(succ, dtype=torch.float)
    data_label = torch.tensor(data_label, dtype=torch.float)
    tr = torch.tensor(tr, dtype=torch.float)
    rssi = data_normal(rssi)

    ff = rssi.numpy()

    for i in range(lens):
        if i >= TIME_STEP:
            # 取时间戳为10，十个一组
            rssi1 = rssi[i - TIME_STEP:i]
            etx1 = etx[i - TIME_STEP:i]
            succ1 = succ[i - TIME_STEP:i]
            label1 = data_label[i - TIME_STEP:i]
            tr1 = tr[i - TIME_STEP:i]
            ff1 = fft.fft(ff[i - TIME_STEP:i])
            ff2 = torch.tensor([f.real for f in ff1], dtype=torch.float)
            ff3 = torch.tensor([f.imag for f in ff1], dtype=torch.float)

            pos.append(postion[i - TIME_STEP:i])
            data_all = torch.stack([rssi1, etx1, succ1, ff2, ff3, tr1], 0)
            # data_all=torch.stack([rssi1,etx1,succ1],0)
            if i == TIME_STEP:
                data = data_all.unsqueeze(dim=0)
            else:
                data = torch.tensor(torch.cat([data, data_all.unsqueeze(dim=0)], 0))

            if (torch.sum(label1) >= 5):
                label.append(1)
            else:
                label.append(0)

    data = torch.tensor(data, dtype=torch.float)
    label = torch.tensor(label, dtype=torch.float)

    return data, label, pos


def nodedistance(point1, point2):
    AB = np.linalg.norm(point1 - point2)
    return AB


# def find_neighbor(j, x, eps):
#     N = list()
#     for i in range(x.shape[0]):
#         temp = np.sqrt(np.sum(np.square(x[j] - x[i])))  # 计算欧式距离
#         #如果距离小于eps
#         if temp <= eps:
#             #append用于在列表末尾添加新的对象
#             N.append(i)
#     #返回邻居的索引
#     return set(N)

# def DBSCAN(X, eps, min_Pts):
#     k = -1
#     neighbor_list = []  # 用来保存每个数据的邻域
#     omega_list = []  # 核心对象集合
#     gama = set([x for x in range(len(X))])  # 初始时将所有点标记为未访问
#     cluster = [-1 for _ in range(len(X))]  # 聚类
#     for i in range(len(X)):
#         neighbor_list.append(find_neighbor(i, X, eps))
#         #取倒数第一个进行if，如果大于设定的样本数，即为核心点
#         if len(neighbor_list[-1]) >= min_Pts:
#             omega_list.append(i)  # 将样本加入核心对象集合
#     omega_list = set(omega_list)  # 转化为集合便于操作
#     while len(omega_list) > 0:
#         #深复制gama
#         gama_old = copy.deepcopy(gama)
#         j = random.choice(list(omega_list))  # 随机选取一个核心对象
#         #k计数，从0开始为第一个
#         k = k + 1
#         #初始化Q
#         Q = list()
#         #记录访问点
#         Q.append(j)
#         #从gama中移除j,剩余未访问点
#         gama.remove(j)
#         while len(Q) > 0:
#             #将第一个点赋值给q,Q队列输出给q,先入先出。
#             q = Q[0]
#             Q.remove(q)
#             if len(neighbor_list[q]) >= min_Pts:
#                 #&按位与运算符：参与运算的两个值,如果两个相应位都为1,则该位的结果为1,否则为0
#                 delta = neighbor_list[q] & gama
#                 deltalist = list(delta)
#                 for i in range(len(delta)):
#                     #在Q中增加访问点
#                     Q.append(deltalist[i])
#                     #从gama中移除访问点,剩余未访问点
#                     gama = gama - delta
#         #原始未访问点-剩余未访问点=访问点
#         Ck = gama_old - gama
#         Cklist = list(Ck)
#         for i in range(len(Ck)):
#             #类型为k
#             cluster[Cklist[i]] = k
#         #剩余核心点
#         omega_list = omega_list - Ck
#     return cluster

# def linear_model(x):
#     return torch.mul(x, w) + b

# def get_loss(my_pred, my_y_train):
#     return torch.mean((my_pred - my_y_train) ** 2)

def GeneralEquation(first_x, first_y, second_x, second_y):
    # 一般式 Ax+By+C=0
    A = second_y - first_y
    B = first_x - second_x
    C = second_x * first_y - first_x * second_y
    k = -1 * A / B
    b = -1 * C / B
    return k, b


# 根据已知两点坐标，求过这两点的直线解析方程： a*x+b*y+c = 0  (a >= 0)
def getLinearEquation(p1x, p1y, p2x, p2y):
    sign = 1
    a = p2y - p1y
    if a < 0:
        sign = -1
        a = sign * a
    b = sign * (p1x - p2x)
    c = sign * (p1y * p2x - p1x * p2y)
    # return [a, b, c]
    # y=kx+bb
    k = -a / b
    bb = -c / b
    return k, bb


def findtwonode(pk, pb):
    pd = [[0, 0], [0, 0], [0, 0], [0, 0]]
    suoying = [0, 0, 0, 0]
    if pb <= 0 and pb >= -480:
        suoying[0] = 1
        pd[0] = [0, pb]
    if (pk * 640 + pb) <= 0 and (pk * 640 + pb) >= -480:
        suoying[1] = 1
        pd[1] = [640, pk * 640 + pb]
    if ((0 - pb) / pk) >= 0 and ((0 - pb) / pk) <= 640:
        suoying[2] = 1
        pd[2] = [(0 - pb) / pk, 0]
    if ((-480 - pb) / pk) >= 0 and ((-480 - pb) / pk) <= 640:
        suoying[3] = 1
        pd[3] = [(-480 - pb) / pk, -480]
    sy = []
    for i in range(len(suoying)):
        if suoying[i] == 1:
            sy.append(i)
    return pd[sy[0]], pd[sy[1]]


def PSM(ltest, ltrue):
    test_x, test_y = findtwonode(ltest[0], ltest[1])
    test_xy = [(test_x[0] + test_y[0]) / 2, (test_x[1] + test_y[1]) / 2]
    true_x, true_y = findtwonode(ltrue[0], ltrue[1])
    true_xy = [(true_x[0] + true_y[0]) / 2, (true_x[1] + true_y[1]) / 2]
    dirt_len = nodedistance(np.array(test_xy), np.array(true_xy))
    return max(1 - (dirt_len / (math.sqrt(480 * 640))), 0)


def LnSM(ltest, ltrue):
    test_x, test_y = findtwonode(ltest[0], ltest[1])
    test_len = nodedistance(np.array(test_x), np.array(test_y))
    true_x, true_y = findtwonode(ltrue[0], ltrue[1])
    true_len = nodedistance(np.array(true_x), np.array(true_y))
    return min(test_len, true_len) / max(test_len, true_len)


def GetCrossAngle(line1):
    [x1, y1, x2, y2] = [0, line1[1], 640, line1[0] * 640 + line1[1]]
    [x3, y3, x4, y4] = [0, 0, 640, 0]
    arr_0 = np.array([(x2 - x1), (y2 - y1)])
    arr_1 = np.array([(x4 - x3), (y4 - y3)])
    cos_value = (float(arr_0.dot(arr_1)) / (np.sqrt(arr_0.dot(arr_0)) * np.sqrt(arr_1.dot(arr_1))))
    if cos_value > 1:
        cos_value = 1
    elif cos_value < -1:
        cos_value = -1
    return np.arccos(cos_value)


def GetClockAngle(line1):
    v1 = [640, line1[0] * 640]
    v2 = [640, 0]
    # 2个向量模的乘积
    TheNorm = np.linalg.norm(v1) * np.linalg.norm(v2)
    # 叉乘
    rho = np.rad2deg(np.arcsin(np.cross(v1, v2) / TheNorm))
    # 点乘
    theta = np.rad2deg(np.arccos(np.dot(v1, v2) / TheNorm))
    if rho < 0:
        return - theta
    else:
        return theta


def OSM(ltest, ltrue):
    # test_angle = GetCrossAngle(ltest)
    # true_angle = GetCrossAngle(ltrue)
    test_angle = GetClockAngle(ltest)
    true_angle = GetClockAngle(ltrue)
    # return max(1-(abs(test_angle-true_angle))/(np.pi/8),0)
    osm = (abs(test_angle - true_angle))
    if osm > 90:
        print('两直线夹角', osm, '\n')

    return max(1 - (abs(test_angle - true_angle)) / (30), 0)


def calculation(ltest, ltrue):
    # LSM = (PSM(ltest,ltrue)+LnSM(ltest,ltrue)+OSM(ltest,ltrue))/3
    LSM = 0.2 * PSM(ltest, ltrue) + 0.8 * OSM(ltest, ltrue)
    return LSM


# def panduan(pr_point,tr_point):
#     if pow(pow(tr_point[0]-pr_point[0],2)+pow(tr_point[1]-pr_point[1],2),0.5)<=35:
#         return 1
#     else:
#         return 0

# result1,result2=[],[]
# for u in range(50):
#     print('number ',u,' start:')
#     line_kb1,line_kb2=[],[]
#     result_pre_send,result_pre_rece = [],[]
#     for v in range(1,3):
#         line_k,line_b=[],[]
#         for i in range(v,TRUE_NUM+1,2):
# data_file ='D:\\GCN\\定位校准\\lastdata\\truedata'+str(i)+'.xlsx'

baocun = [[1, 'D:\\GCN\\定位校准\\lastdata\\测试\\nodedata3.xlsx', 63, -131, 553, -322],
          [2, 'D:\\GCN\\定位校准\\lastdata\\测试\\nodedata3.xlsx', 63, -131, 553, -322],
          [3, 'D:\\GCN\\定位校准\\lastdata\\测试\\nodedata1.xlsx', 90, -127, 594, -317],
          [4, 'D:\\GCN\\定位校准\\lastdata\\定位\\nodedata3.xlsx', 101, -138, 42, -406],
          [5, 'D:\\GCN\\定位校准\\lastdata\\定位\\nodedata3.xlsx', 101, -138, 42, -406],
          [6, 'D:\\GCN\\定位校准\\lastdata\\定位\\nodedata4.xlsx', 101, -137, 251, -406],
          [7, 'D:\\GCN\\定位校准\\lastdata\\定位\\nodedata5.xlsx', 218, -111, 42, -410],
          [8, 'D:\\GCN\\定位校准\\lastdata\\定位\\nodedata7.xlsx', 379, -112, 41, -411],
          [9, 'D:\\GCN\\定位校准\\lastdata\\定位\\nodedata8.xlsx', 379, -112, 256, -408],
          [10, 'D:\\GCN\\定位校准\\lastdata\\定位\\nodedata9.xlsx', 544, -139, 41, -409],
          [11, 'D:\\GCN\\定位校准\\lastdata\\定位\\nodedata10.xlsx', 541, -188, 254, -382],
          [12, 'D:\\GCN\\定位校准\\lastdata\\定位\\nodedata11.xlsx', 638, -252, 42, -410],
          [13, 'D:\\GCN\\定位校准\\lastdata\\定位\\nodedata11.xlsx', 638, -252, 42, -410],
          [14, 'D:\\GCN\\定位校准\\lastdata\\走廊\\nodedata1.xlsx', 466, -153, 470, -361],
          [15, 'D:\\GCN\\定位校准\\lastdata\\走廊\\nodedata3.xlsx', 469, -120, 628, -318],
          [16, 'D:\\GCN\\定位校准\\lastdata\\走廊\\nodedata3.xlsx', 469, -120, 628, -318],
          [17, 'D:\\GCN\\定位校准\\lastdata\\走廊\\nodedata4.xlsx', 173, -213, 629, -315],
          [18, 'D:\\GCN\\定位校准\\lastdata\\走廊\\nodedata5.xlsx', 486, -133, 194, -309],
          [19, 'D:\\GCN\\定位校准\\lastdata\\排球场\\nodedata3.xlsx', 490, -434, 484, -239],
          [20, 'D:\\GCN\\定位校准\\lastdata\\排球场\\nodedata3.xlsx', 490, -434, 484, -239],
          [21, 'D:\\GCN\\定位校准\\lastdata\\排球场\\nodedata3.xlsx', 490, -434, 484, -239],
          [22, 'D:\\GCN\\定位校准\\lastdata\\排球场\\nodedata3.xlsx', 490, -434, 484, -239],
          [23, 'D:\\GCN\\定位校准\\lastdata\\排球场\\nodedata4.xlsx', 67, -272, 464, -389],
          [24, 'D:\\GCN\\定位校准\\lastdata\\排球场\\nodedata5.xlsx', 504, -314, 317, -178],
          [25, 'D:\\GCN\\定位校准\\lastdata\\排球场\\nodedata6.xlsx', 33, -351, 317, -179],
          [26, 'D:\\GCN\\定位校准\\lastdata\\大厅\\nodedata2.xlsx', 184, -156, 369, -380],
          [27, 'D:\\GCN\\定位校准\\lastdata\\会议室\\nodedata2.xlsx', 25, -346, 507, -357],
          [28, 'D:\\GCN\\定位校准\\lastdata\\会议室\\nodedata2.xlsx', 25, -346, 507, -357],
          [29, 'D:\\GCN\\定位校准\\lastdata\\会议室\\nodedata10.xlsx', 105, -270, 330, -307],
          [30, 'D:\\GCN\\定位校准\\lastdata\\会议室\\nodedata9.xlsx', 562, -252, 330, -307],
          [31, 'D:\\GCN\\定位校准\\lastdata\\会议室\\nodedata12.xlsx', 90, -318, 451, -311],
          [32, 'D:\\GCN\\定位校准\\lastdata\\会议室\\nodedata12.xlsx', 90, -318, 451, -311],
          [33, 'D:\\GCN\\graphnet\\data\\室内定位1\\30-1\\data1.xlsx', 332, -228, 637, -213],
          [34, 'D:\\GCN\\graphnet\\data\\室内定位1\\30-2\\data2.xlsx', 312, -218, 638, -227],
          [35, 'D:\\GCN\\graphnet\\data\\室内定位1\\30-2\\data3.xlsx', 77, -68, 638, -232],
          [36, 'D:\\GCN\\定位校准\\lastdata\\定位\\nodedata8.xlsx', 379, -112, 256, -408],
          [37, 'D:\\GCN\\定位校准\\lastdata\\定位\\nodedata9.xlsx', 544, -139, 41, -409],
          [38, 'D:\\GCN\\定位校准\\lastdata\\排球场\\nodedata3.xlsx', 490, -434, 484, -239],
          [39, 'D:\\GCN\\定位校准\\lastdata\\大厅\\nodedata2.xlsx', 184, -156, 369, -380],
          [40, 'D:\\GCN\\定位校准\\lastdata\\大厅\\nodedata2.xlsx', 184, -156, 369, -380]
          ]

for sumid in range(len(baocun)):
    # for sumid in range(10):
    for itemvv in number_id:
        data_file = baocun[sumid][1]
        true_data, true_label, true_postion = load_data(data_file)
        true_data = np.array(true_data)
        true_label = np.array(true_label)
        pos_num_dict = {}
        pos_sum_dict = {}
        pos_dict = {}
        for items in true_postion:
            for key in items:
                pos_num_dict[key] = pos_num_dict.get(key, 0) + 1
                pos_sum_dict[key] = 0

        clf = load_learner("D:\\GCN\\定位校准\\models\\md.pkl") # 加载射频遮挡检测模型
        probas, target, preds = clf.get_X_preds(true_data, true_label, with_decoded=True)

        preds = [int(float(x)) for x in preds]
        # 计算模型指标
        acc = accuracy_score(target, preds)
        pre = precision_score(target, preds)
        rec = recall_score(target, preds)
        fpr, tpr, thresholds_keras = roc_curve(target, preds)
        auc_score = auc(fpr, tpr)
        print(f'Accuracy on : ', acc)
        print(f'Pre on : ', pre)
        print(f'Rec on : ', rec)
        print(f'FPR on : ', fpr)
        print(f'TPR on : ', tpr)
        print(f'AUC on : ', auc_score)

        for j in range(len(preds)):
            for key in true_postion[j]:
                pos_sum_dict[key] += int(float(preds[j]))

        for key in pos_num_dict:
            pos_dict[key] = pos_sum_dict[key] / pos_num_dict[key]

        x0, y0, z0 = [], [], []
        for key in pos_dict:
            if pos_dict[key] > 0.7:
                # for count in range(pos_num_dict[key]):
                x0.append(key[0])
                y0.append(-key[1])
                z0.append(pos_dict[key])

        X = [[j, k, l] for j, k, l in zip(x0, y0, z0)]
        Y = []
        for i in range(400):
            index_old = [q for q in range(len(X))]
            # 随机取30%的点
            # index_new = random.sample(index_old,int(0.02*len(X)))
            # 随机取10个点
            index_new = random.sample(index_old, itemvv)
            Y.append(np.array([X[item] for item in index_new]))
            # plt.xlim(0,640)
            # plt.ylim(-480,0)
            # plt.plot(Y[i][:,0],Y[i][:,1], 'o')
            # plt.show()
            # print()
        Z = []
        for i in range(len(Y)):
            distance = []
            for j in range(len(Y[i])):
                for k in range(len(Y[i])):
                    distance.append(nodedistance(Y[i][j], Y[i][k]))
            distance = np.array(distance)
            Z.append(np.std(distance))
        # # # # print(Z)
        if sumid == 5 or sumid == 8 or sumid == 17 or sumid == 18 or sumid == 35 or sumid == 37:
            l1, l2 = (list(t) for t in zip(*sorted(zip(Z, Y), reverse=False)))
        else:
            l1, l2 = (list(t) for t in zip(*sorted(zip(Z, Y), reverse=True)))
        line_k, line_b = [], []
        clf = sklearn.linear_model.Ridge(alpha=0.5)
        for i in range(100):
            print('number: ', i)

            x1, y1 = [], []
            for j in range(len(l2[i])):
                # x1.append(Y[i][j][0])
                # y1.append(Y[i][j][1])
                x1.append(l2[i][j][0])
                y1.append(l2[i][j][1])

            pointx = x1
            pointy = y1
            # 导入数据集
            px = np.array(pointx).reshape(-1, 1)
            # px = np.array(pointx)
            py = np.array(pointy)

            clf.fit(px, py)
            y_pred = clf.predict(px)
            MSN1 = mean_squared_error(y_pred, py)
            for u in range(len(y_pred)):
                if (px[u] != px[-1]):
                    break
            k1, b1 = GeneralEquation(px.reshape(1, -1).squeeze()[u], y_pred[u], px.reshape(1, -1).squeeze()[-1],
                                     y_pred[-1])

            px = np.array([-i for i in pointy]).reshape(-1, 1)
            py = np.array([-i for i in pointx])

            clf.fit(px, py)
            y_pred = clf.predict(px)
            MSN2 = mean_squared_error(y_pred, py)

            p_y = np.array([-i for i in px.reshape(1, -1).squeeze()])
            p_x = np.array([-i for i in py])
            ppx = np.array([-i for i in y_pred])
            for n in range(len(p_y)):
                if (ppx[n] != ppx[-1]):
                    break
            k2, b2 = GeneralEquation(ppx.reshape(1, -1).squeeze()[n], p_y[n], ppx.reshape(1, -1).squeeze()[-1], p_y[-1])
            if MSN1 < MSN2:
                line_k.append(k1)
                line_b.append(b1)
            else:
                line_k.append(k2)
                line_b.append(b2)
        line_test = [[k, b] for k, b in zip(line_k, line_b)]

        k_true, b_true = getLinearEquation(baocun[sumid][2], baocun[sumid][3], baocun[sumid][4], baocun[sumid][5])
        line_true = [float(k_true), float(b_true)]

        label = []
        data1, data2 = [], []
        for i in range(len(line_test)):
            result = calculation(line_test[i], line_true)
            label.append(result)
            data1 = [label[i]] + line_true + line_test[i] + l2[i].tolist()
            data2.append(data1)
        print(data2)

        m = np.array(data2)
        np.save('D:\\GCN\\定位校准\\dingwei_data\\data5\\' + str(itemvv) + '\\' + str(sumid + 1) + '_' + str(
            itemvv) + '.npy', m)
