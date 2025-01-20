'''

'''
import torch
from torch.utils.data import DataLoader
import torch.optim as optim
from time import time
from tqdm import tqdm
import os
from sklearn.metrics import roc_curve
from sklearn.metrics import auc
from sklearn.metrics import recall_score,accuracy_score
from sklearn.metrics import precision_score,f1_score

from torch.utils.data import TensorDataset
import openpyxl 
from openpyxl import load_workbook 
import csv
from scipy import fft
import numpy as np
from sklearn.cluster import KMeans
import itertools
import datetime
import bisect

def load_camera(path):
    wb_camera = load_workbook(path)
    sh_camera = wb_camera['Sheet']
    time = []
    obstacle = []
    for i in range(1,sh_camera.max_row+1):
        time.append(datetime.datetime.strptime(sh_camera.cell(i,1).value.replace('\r','').replace('\n','').replace('\t',''), '%Y-%m-%d-%H:%M:%S.%f'))
        obs = []
        # for j in range(2,26,5):
        for j in range(2,26,5):
            # if int(sh_camera.cell(i,j).value) == -1:
            #     if j==2:
            #         obs.append([-1,-1,-1,-1,-1])
            #         break
            #     else:
            #         break
            # else:
            obs.append([float(sh_camera.cell(i, j).value),float(sh_camera.cell(i,j+1).value),float(sh_camera.cell(i,j+2).value),float(sh_camera.cell(i,j+3).value),float(sh_camera.cell(i,j+4).value)])
        obstacle.append(obs)
    obstacle_dict = dict(zip(time, obstacle))
    return time, obstacle_dict

def load_linenode(path):
    wb_node = load_workbook(path)
    sh_node = wb_node['Sheet']
    line_node = []
    line_time = []
    first_rssi = float(sh_node.cell(1,3).value)
    for i in range(1,sh_node.max_row+1):
        time = datetime.datetime.strptime(sh_node.cell(i,1).value.replace('\r','').replace('\n','').replace('\t',''), '%Y-%m-%d-%H:%M:%S.%f')
        line_node.append([int(sh_node.cell(i, 2).value),  # 序号
                          float(sh_node.cell(i, 3).value),  # rssi
                          first_rssi-(float(sh_node.cell(i, 3).value)),  # rssi[0] - rssi[now]
                          float(sh_node.cell(i, 4).value),  # etx:[0, 1]
                          int(sh_node.cell(i, 8).value)])  # label
        line_time.append(time)
    return line_time, line_node

scene = [4]
line_number = 4

if __name__ == '__main__':
    for sc in scene:
        camera_file ='D:\\GCN\\graphnet\\data\\train_data\\outdoor\\30-'+str(sc)+'\\t_camera.xlsx'
        obstacle_index, obstacle = load_camera(camera_file)
        linenode = []
        linetime = []
        for li in range(line_number):
            linenode_file ='D:\\GCN\\graphnet\\data\\train_data\\outdoor\\30-'+str(sc)+'\\data'+str(li+1)+'.xlsx'
            lt, ln = load_linenode(linenode_file)
            linetime.append(lt)
            linenode.append(ln)
        
        last_list = [linenode[:][i][-1][0] for i in range(len(linenode))]
        last_time = min(last_list)
        last_index = last_list.index(last_time)
        node_feature = []
        node_label = []
        obstacle_feature = []
        node_time = []
        for i in range(len(linenode[last_index])):
            time_list = [linetime[:][k][i] for k in range(len(linetime))]
            time_key = max(time_list)
            no_feature = []
            no_label = []
            for j in range(len(linenode)):
                no_feature.append(linenode[j][i][1:4])
                no_label.append(linenode[j][i][-1])
            
            left_index = bisect.bisect_left(obstacle_index, time_key)
            right_index = bisect.bisect_right(obstacle_index, time_key)
            if left_index == right_index:
                best_index = left_index
            else:
                if abs(obstacle_index[left_index] - time_key) < abs(obstacle_index[right_index] - time_key):
                    best_index = left_index
                else:
                    best_index = right_index
                    
            ob_feature = obstacle.get(obstacle_index[best_index])
            node_time.append(obstacle_index[best_index])
            node_feature.append(no_feature)
            node_label.append(no_label)
            obstacle_feature.append(ob_feature)

        np.save('D:\\GCN\\graphnet\\data\\train_data\\outdoor\\30-'+str(sc)+'\\dataset\\time.npy',np.array(node_time))
        np.save('D:\\GCN\\graphnet\\data\\train_data\\outdoor\\30-'+str(sc)+'\\dataset\\feature.npy',np.array(node_feature))
        np.save('D:\\GCN\\graphnet\\data\\train_data\\outdoor\\30-'+str(sc)+'\\dataset\\label.npy',np.array(node_label))
        np.save('D:\\GCN\\graphnet\\data\\train_data\\outdoor\\30-'+str(sc)+'\\dataset\\obstacle.npy',np.array(obstacle_feature))
    print()